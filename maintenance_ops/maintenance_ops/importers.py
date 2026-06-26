from __future__ import annotations

from pathlib import Path

import csv
from openpyxl import load_workbook

from .api import normalize_city_code
from .models import ExceptionRecord, Outlet, TeamMember


def import_outlets(xlsx_path: str) -> tuple[list[Outlet], list[ExceptionRecord]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    records: list[Outlet] = []
    exceptions: list[ExceptionRecord] = []

    for idx, row in enumerate(rows[1:], start=2):
        city, outlet_code = row[:2]
        if not city or not outlet_code:
            exceptions.append(ExceptionRecord(
                source_file=Path(xlsx_path).name,
                source_row_number=idx,
                issue_type="MissingOutletData",
                issue_message="Missing city or outlet code",
                source_payload=str(row),
            ))
            continue
        records.append(Outlet(
            outlet_code=str(outlet_code).strip(),
            outlet_name=str(outlet_code).strip(),
            city=normalize_city_code(str(city)),
        ))
    return records, exceptions


def import_team_members(csv_path: str) -> tuple[list[TeamMember], list[ExceptionRecord]]:
    records: list[TeamMember] = []
    exceptions: list[ExceptionRecord] = []
    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for idx, row in enumerate(reader, start=2):
            if not row.get("Employee No") or not row.get("Name"):
                exceptions.append(ExceptionRecord(
                    source_file=Path(csv_path).name,
                    source_row_number=idx,
                    issue_type="MissingTeamData",
                    issue_message="Missing employee number or name",
                    source_payload=str(row),
                ))
                continue
            records.append(TeamMember(
                employee_no=row["Employee No"].strip(),
                employee_name=row["Name"].strip(),
                job_title=row.get("Job title", "").strip(),
                department=row.get("Department", "").strip(),
                email=row.get("Email", "").strip(),
                mobile=row.get("Mobile", "").strip(),
                reports_to=row.get("Reports to", "").strip() or None,
                home_office=row.get("Home", "").strip() or None,
            ))
    return records, exceptions


def extract_ticket_taxonomy(xlsx_path: str) -> tuple[list[dict], list[ExceptionRecord]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    records: list[dict] = []
    exceptions: list[ExceptionRecord] = []
    for idx, row in enumerate(rows[1:], start=2):
        department, category, sub1, sub2 = row[:4]
        if not department or not category:
            exceptions.append(ExceptionRecord(
                source_file=Path(xlsx_path).name,
                source_row_number=idx,
                issue_type="MissingTaxonomyData",
                issue_message="Missing department or category",
                source_payload=str(row),
            ))
            continue
        records.append({
            "department": str(department).strip(),
            "category": str(category).strip(),
            "sub_category_1": (str(sub1).strip() if sub1 else ""),
            "sub_category_2": (str(sub2).strip() if sub2 else ""),
        })
    return records, exceptions


def extract_pm_tracker(xlsx_path: str) -> tuple[list[dict], list[ExceptionRecord]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    records: list[dict] = []
    exceptions: list[ExceptionRecord] = []

    for idx, row in enumerate(rows[1:], start=2):
        outlet_code, city, asset, task, freq = row[:5]
        if not outlet_code or not asset:
            exceptions.append(
                ExceptionRecord(
                    source_file=Path(xlsx_path).name,
                    source_row_number=idx,
                    issue_type="MissingPMData",
                    issue_message="Missing outlet or asset in PM tracker",
                    source_payload=str(row),
                )
            )
            continue

        records.append(
            {
                "outlet_code": str(outlet_code).strip(),
                "city": normalize_city_code(str(city)) if city else "",
                "asset_category": str(asset).strip(),
                "task": str(task).strip() if task else "",
                "frequency": str(freq).strip() if freq else "",
            }
        )

    return records, exceptions
